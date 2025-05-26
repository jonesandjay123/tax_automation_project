import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import google.generativeai as genai # Added for Gemini API

# API Key for Gemini - Ensure this is kept secure and not hardcoded in production
API_KEY = "AIzaSyCBopfgQ5FFcKvvy52_pUAmZ33I6DSoAdI"

# Configure the Gemini client globally
try:
    genai.configure(api_key=API_KEY)
except Exception as e:
    print(f"CRITICAL: Error configuring Gemini API: {e}. Summarization will be skipped.")
    # Set genai to None or a flag to indicate it's not available
    genai = None 

def call_gemini_to_summarize(text_to_summarize, base_name):
    """
    Calls Gemini API to summarize the given text.
    Returns summarized text or original text if API call fails.
    """
    if not genai:
        print(f"Skipping Gemini summary for {base_name} as API is not configured.")
        return f"[Gemini disabled] {text_to_summarize[:150]}..." # Return truncated original if Gemini is off
    
    print(f"[Gemini] Summarizing content for: {base_name}...")
    try:
        model = genai.GenerativeModel('gemini-1.5-flash-latest') # Using a generally available model
        # Simple prompt for summarization, can be made more sophisticated
        prompt = f"Summarize the following tax information for '{base_name}' in one concise sentence, focusing on the tax rate and key conditions: \n{text_to_summarize}"
        response = model.generate_content(prompt)
        if response.text:
            print(f"[Gemini] Successfully summarized: {base_name}")
            return response.text.strip()
        else:
            print(f"[Gemini] Warning: Empty response for {base_name}. Returning original text.")
            return text_to_summarize
    except Exception as e:
        print(f"[Gemini] Error summarizing {base_name}: {e}. Returning original text.")
        return text_to_summarize # Fallback to original text

def scrape_ny_tax_info():
    """
    Scrapes tax information from the NYS tax website.
    Returns a dictionary with extracted tax base information or None if an error occurs.
    """
    url = "https://www.tax.ny.gov/bus/ct/def_art9a.htm"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching URL: {e}")
        return None

    soup = BeautifulSoup(response.content, "html.parser")
    tax_bases_data = {}

    def get_text_after_heading(heading_text, next_heading_text=None):
        # Try finding by specific IDs first for more robustness if available
        if "Entire Net Income" in heading_text:
            heading_tag = soup.find(id="eni")
            if heading_tag:
                 # The actual content might be under the next h3/h4 after the anchor
                 heading_tag = heading_tag.find_next(["h3","h4"])
        else:
            heading_tag = soup.find(lambda tag: tag.name in ["h3", "h4", "strong"] and heading_text in tag.get_text())
        
        if not heading_tag:
            print(f"Could not find heading for: {heading_text}")
            return "Not found or website structure changed."

        content = []
        current_tag = heading_tag.find_next_sibling()
        limit_tags = ["h2", "h3", "h4"]

        # Custom stop conditions for each section if necessary
        stop_texts_eni = ["Business capital base", "Fixed dollar minimum tax"]
        stop_texts_capital = ["Fixed dollar minimum tax", "Metropolitan transportation business tax"]
        stop_texts_fdm = ["Metropolitan transportation business tax", "S corporations"] # Adjust if needed

        current_stop_texts = []
        if "Entire Net Income" in heading_text:
            current_stop_texts = stop_texts_eni
        elif "Business capital base" in heading_text:
            current_stop_texts = stop_texts_capital
        elif "Fixed dollar minimum tax" in heading_text:
            current_stop_texts = stop_texts_fdm

        processed_tags_count = 0
        max_tags_to_process = 10 # Limit to avoid infinite loops on structure changes

        while current_tag and processed_tags_count < max_tags_to_process:
            processed_tags_count += 1
            if current_tag.name in limit_tags:
                tag_text = current_tag.get_text(strip=True)
                # Check if this is a stop-heading for the current section
                if any(stop_text in tag_text for stop_text in current_stop_texts):
                    break
                # If it's a heading but not a stop heading for this section, 
                # and not the original heading, it might be an unexpected new section.
                if heading_text not in tag_text: 
                    is_target_keyword_present = False
                    for keyword in ["Entire Net Income", "Capital Base", "Fixed Dollar Minimum"]:
                        if keyword in tag_text: # if it's one of our other target sections, then we should stop
                            is_target_keyword_present = True
                            break
                    if is_target_keyword_present: # if it is another target keyword, then break
                        break
                    # If it is not a target keyword and is not the original heading, it is a sub heading, continue parsing
            
            text_content = ""
            if current_tag.name == 'p':
                text_content = current_tag.get_text(strip=True, separator=" ")
            elif current_tag.name == 'ul':
                list_items = []
                for li in current_tag.find_all('li'):
                    li_text = li.get_text(strip=True, separator=" ")
                    if li_text:
                        list_items.append(li_text)
                text_content = "; ".join(list_items)
            
            if text_content:
                content.append(text_content)
            
            # Limit paragraphs/elements to avoid overly long content for one base
            if len(content) > 3 and "Fixed dollar minimum tax" not in heading_text:
                break 
            if len(content) > 6: # More content for FDM as it has tables
                break

            current_tag = current_tag.find_next_sibling()
            
        full_content = " ".join(content) if content else "Content not found or structure changed."
        return full_content

    # Extracting texts
    raw_eni_text = get_text_after_heading("Entire Net Income (ENI) base")
    raw_capital_text = get_text_after_heading("Business capital base")
    raw_fdm_text = get_text_after_heading("Fixed dollar minimum tax")

    # Store raw data first
    tax_bases_data["Entire Net Income Base"] = raw_eni_text
    tax_bases_data["Capital Base"] = raw_capital_text
    tax_bases_data["Fixed Dollar Minimum"] = raw_fdm_text
    
    # Attempt to refine/summarize extracted text using Gemini
    # The prompt to Gemini will ask for rates and conditions.

    summarized_tax_bases_data = {}
    if raw_eni_text and raw_eni_text != "Not found or website structure changed.":
        summarized_tax_bases_data["Entire Net Income Base"] = call_gemini_to_summarize(raw_eni_text, "Entire Net Income Base")
    else:
        summarized_tax_bases_data["Entire Net Income Base"] = raw_eni_text # Keep as is if not found

    if raw_capital_text and raw_capital_text != "Not found or website structure changed.":
        summarized_tax_bases_data["Capital Base"] = call_gemini_to_summarize(raw_capital_text, "Capital Base")
    else:
        summarized_tax_bases_data["Capital Base"] = raw_capital_text

    if raw_fdm_text and raw_fdm_text != "Not found or website structure changed.":
        summarized_tax_bases_data["Fixed Dollar Minimum"] = call_gemini_to_summarize(raw_fdm_text, "Fixed Dollar Minimum Tax")
    else:
        summarized_tax_bases_data["Fixed Dollar Minimum"] = raw_fdm_text
        
    return summarized_tax_bases_data

def create_excel_summary(tax_data):
    """
    Creates an Excel file with the tax summary.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "NY Tax Summary"

    headers = {
        'A1': "State", 'B1': "Nexus Standard", 'C1': "Effective Date (Nexus)",
        'D1': "Tax Base Summary (Gemini Summarized if available)", 
        'E1': "Tax Rates (included in D1)", 'F1': "Source URL",
        'G1': "Sales Factor Method", 'H1': "Effective Date (Sales Factor)"
    }
    for cell, header in headers.items():
        ws[cell] = header

    if tax_data:
        summary_string = f"Entire Net Income Base: {tax_data.get('Entire Net Income Base', 'Not found')}; " \
                         f"Capital Base: {tax_data.get('Capital Base', 'Not found')}; " \
                         f"Fixed Dollar Minimum: {tax_data.get('Fixed Dollar Minimum', 'Not found')}"
    else:
        summary_string = "Could not retrieve or summarize tax data."

    ws['A2'] = "new york"
    ws['B2'] = "market base"
    ws['C2'] = "2014"
    ws['D2'] = summary_string
    ws['F2'] = "https://www.tax.ny.gov/bus/ct/def_art9a.htm#eni"
    ws['G2'] = "market base"
    ws['H2'] = "2014"

    output_dir = "output"
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except OSError as e:
            print(f"Error creating directory {output_dir}: {e}")
            return # Stop if directory cannot be created
    
    filepath = os.path.join(output_dir, "ny_tax_summary.xlsx")
    
    try:
        wb.save(filepath)
        print(f"Excel file '{filepath}' created successfully.")
    except PermissionError:
        print(f"ERROR: Permission denied when trying to save '{filepath}'. Please ensure the file is not open in another program (e.g., Excel) and that you have write permissions to the directory.")
    except Exception as e:
        print(f"Error saving Excel file '{filepath}': {e}")

if __name__ == "__main__":
    # Scrape data (this will now include Gemini summarization attempts)
    summarized_data = scrape_ny_tax_info()
    
    if summarized_data:
        print("\n--- Summarized Tax Info (from Gemini or fallback) ---")
        for base, info in summarized_data.items():
            print(f"{base}: {info}")
        print("----------------------------------------------------\n")
    else:
        print("Failed to scrape data. Excel file will contain placeholders or errors.")
    
    create_excel_summary(summarized_data) 
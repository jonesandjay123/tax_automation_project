import openpyxl
from openpyxl import Workbook
import os
import google.generativeai as genai # Added for Gemini API

# API Key for Gemini
API_KEY = "AIzaSyCBopfgQ5FFcKvvy52_pUAmZ33I6DSoAdI"

# Configure the Gemini client
try:
    genai.configure(api_key=API_KEY)
except Exception as e:
    print(f"Error configuring Gemini API: {e}")
    # Optionally, exit or handle more gracefully
    exit()

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Sample Data"

# Add some headers
headers = ["Name", "Age", "City"]
ws.append(headers)

# Add some sample data
data = [
    ("Alice", 30, "New York"),
    ("Bob", 24, "San Francisco"),
    ("Charlie", 35, "London"),
    ("David", 28, "Paris")
]

for row in data:
    ws.append(row)

# Call Gemini API and write to Excel
try:
    model = genai.GenerativeModel('gemini-1.5-flash-latest') # Using a standard public model
    prompt = "Explain how AI works in a few words"
    response = model.generate_content(prompt)
    gemini_output = response.text

    ws['A5'] = "Gemini API Response:"
    ws['B5'] = gemini_output
    print("Successfully fetched response from Gemini API and added to Excel.")

except Exception as e:
    print(f"Error calling Gemini API or writing to Excel: {e}")
    ws['A5'] = "Gemini API Response:"
    ws['B5'] = f"Error: {e}"

# Define the output directory and filename
output_dir = "output"
filename = "sample_excel_file.xlsx"
filepath = os.path.join(output_dir, filename)

# Create the output directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Save the workbook
wb.save(filepath)

print(f"Excel file '{filepath}' created successfully.") 